from __future__ import annotations

import io
from typing import Any

from openpyxl import load_workbook

from app.services.import_extractors.base import BaseExtractor, ExtractedTable, ExtractionResult


class XlsxExtractor(BaseExtractor):
    source_type = "xlsx"

    def extract(self, *, filename: str, raw_bytes: bytes, options: dict[str, Any] | None = None) -> ExtractionResult:
        workbook = load_workbook(io.BytesIO(raw_bytes), data_only=True, read_only=True)
        tables: list[ExtractedTable] = []
        total_rows = 0

        for worksheet in workbook.worksheets:
            values = list(worksheet.iter_rows(values_only=True))
            cleaned = [["" if cell is None else str(cell).strip() for cell in row] for row in values]
            cleaned = [row for row in cleaned if any(cell for cell in row)]
            if not cleaned:
                continue

            header_row_idx = self._detect_header_row(cleaned)
            header = cleaned[header_row_idx]
            width = max(len(row) for row in cleaned)
            columns = [self._normalize_header(header[idx] if idx < len(header) else "", idx) for idx in range(width)]
            data_rows = cleaned[header_row_idx + 1 :]
            normalized_rows = [
                {columns[idx]: (row[idx] if idx < len(row) else "") for idx in range(len(columns))}
                for row in data_rows
                if any((row[idx] if idx < len(row) else "") for idx in range(len(columns)))
            ]
            total_rows += len(normalized_rows)
            if normalized_rows:
                tables.append(
                    ExtractedTable(
                        name=worksheet.title,
                        columns=columns,
                        rows=normalized_rows,
                        confidence=0.9,
                        meta={"sheet": worksheet.title, "header_row_index": header_row_idx},
                    )
                )

        return ExtractionResult(source_type=self.source_type, tables=tables, meta={"sheet_count": len(tables), "row_count": total_rows})

    @staticmethod
    def _detect_header_row(rows: list[list[str]]) -> int:
        best_index = 0
        best_score = -1
        for idx, row in enumerate(rows[:10]):
            non_empty = sum(1 for cell in row if cell)
            alpha_cells = sum(1 for cell in row if any(ch.isalpha() for ch in cell))
            score = non_empty * 2 + alpha_cells
            if score > best_score:
                best_score = score
                best_index = idx
        return best_index

    @staticmethod
    def _normalize_header(value: str | None, idx: int) -> str:
        text = (value or "").strip()
        return text or f"column_{idx + 1}"
